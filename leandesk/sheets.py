from __future__ import annotations

import ast
import csv
import io
import json
import uuid
import math
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk
from typing import Any

from .core import RecentFiles, RecoveryRecord, RecoveryStore, atomic_write_json
from .data_boundary import DataCorruptionError, UnsupportedSchemaVersion, merge_known_and_extra, read_bounded, strict_json_load_bytes
from .compatibility import SHEETS_COMPAT, convert_with_libreoffice
from .ui import COLORS, StatusBar, ResponsiveToolbar
from .save_policy import (
    ImportedSourceProtectionError,
    SavePolicyError,
    UnsupportedSaveFormatError,
    imported_source_for,
    mark_save_boundary,
    validate_destination,
    write_atomically,
)

CELL_RE = re.compile(r"\b([A-Z]{1,3}[1-9][0-9]*)\b")
RANGE_RE = re.compile(r"([A-Z]{1,3}[1-9][0-9]*):([A-Z]{1,3}[1-9][0-9]*)")
FUNCTION_RE = re.compile(r"\b(SUM|AVERAGE|MIN|MAX|COUNT|COUNTA|ABS|ROUND)\(([^()]*)\)", re.IGNORECASE)
MAX_ROWS = 200
MAX_COLS = 52
MAX_FORMULA_LENGTH = 4096
MAX_FORMULA_NODES = 128
MAX_FORMULA_DEPTH = 32
MAX_FORMULA_EXPONENT = 100
MAX_FORMULA_MAGNITUDE = 1e100
MAX_FORMULA_REFERENCES = 10_000
WORKBOOK_FORMAT_VERSION = 1
SHEET_CANVAS = "#FFFFFF"
SHEET_TEXT = "#172033"
SHEET_GRID = "#A7B0BE"
SHEET_RANGE_FILL = "#DCE8F7"


def column_name(index: int) -> str:
    result = ""
    value = index + 1
    while value:
        value, remainder = divmod(value - 1, 26)
        result = chr(65 + remainder) + result
    return result


def column_index(name: str) -> int:
    value = 0
    for char in name.upper():
        value = value * 26 + ord(char) - 64
    return value - 1


def split_cell(address: str) -> tuple[int, int]:
    match = re.fullmatch(r"([A-Z]+)([1-9][0-9]*)", address.upper())
    if not match:
        raise ValueError(address)
    return int(match.group(2)) - 1, column_index(match.group(1))


def normalize_cell(address: str) -> str:
    if not isinstance(address, str):
        raise ValueError("Cell address must be text")
    key = address.strip().upper()
    row, col = split_cell(key)
    if row >= MAX_ROWS or col >= MAX_COLS:
        raise ValueError(
            f"Cell {key} is outside LeanDesk's visible grid ({column_name(MAX_COLS - 1)}{MAX_ROWS})."
        )
    return key


def iter_range(start: str, end: str):
    r1, c1 = split_cell(start)
    r2, c2 = split_cell(end)
    for row in range(min(r1, r2), max(r1, r2) + 1):
        for col in range(min(c1, c2), max(c1, c2) + 1):
            yield f"{column_name(col)}{row + 1}"


@dataclass(frozen=True)
class SelectionRange:
    anchor: str = "A1"
    active: str = "A1"

    def bounds(self) -> tuple[int, int, int, int]:
        anchor_row, anchor_col = split_cell(normalize_cell(self.anchor))
        active_row, active_col = split_cell(normalize_cell(self.active))
        return (
            min(anchor_row, active_row),
            min(anchor_col, active_col),
            max(anchor_row, active_row),
            max(anchor_col, active_col),
        )

    def addresses(self) -> list[str]:
        top, left, bottom, right = self.bounds()
        return [
            f"{column_name(col)}{row + 1}"
            for row in range(top, bottom + 1)
            for col in range(left, right + 1)
        ]

    def label(self) -> str:
        top, left, bottom, right = self.bounds()
        start = f"{column_name(left)}{top + 1}"
        end = f"{column_name(right)}{bottom + 1}"
        return start if start == end else f"{start}:{end}"


class SafeMath:
    """Bounded numeric AST evaluator; never calls Python ``eval``."""

    _binops = {
        ast.Add: lambda a, b: a + b,
        ast.Sub: lambda a, b: a - b,
        ast.Mult: lambda a, b: a * b,
        ast.Div: lambda a, b: a / b,
        ast.FloorDiv: lambda a, b: a // b,
        ast.Mod: lambda a, b: a % b,
    }

    def __init__(self) -> None:
        self.nodes = 0

    @staticmethod
    def _bounded(value: float | int) -> float:
        if isinstance(value, bool) or not isinstance(value, (int, float)):
            raise ValueError("Formula did not return a number")
        number = float(value)
        if not math.isfinite(number) or abs(number) > MAX_FORMULA_MAGNITUDE:
            raise ValueError("Formula result exceeds the supported numeric range")
        return number

    def evaluate(self, node: ast.AST, depth: int = 0) -> float:
        self.nodes += 1
        if self.nodes > MAX_FORMULA_NODES or depth > MAX_FORMULA_DEPTH:
            raise ValueError("Formula is too complex")
        if isinstance(node, ast.Expression):
            return self.evaluate(node.body, depth + 1)
        if isinstance(node, ast.Constant):
            return self._bounded(node.value)
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.UAdd, ast.USub)):
            value = self.evaluate(node.operand, depth + 1)
            return self._bounded(value if isinstance(node.op, ast.UAdd) else -value)
        if isinstance(node, ast.BinOp):
            left = self.evaluate(node.left, depth + 1)
            right = self.evaluate(node.right, depth + 1)
            if isinstance(node.op, ast.Pow):
                if abs(right) > MAX_FORMULA_EXPONENT:
                    raise ValueError("Formula exponent exceeds the supported limit")
                if left == 0 and right < 0:
                    raise ZeroDivisionError
                try:
                    return self._bounded(left ** right)
                except (OverflowError, ValueError) as exc:
                    raise ValueError("Formula result exceeds the supported numeric range") from exc
            operation = self._binops.get(type(node.op))
            if operation is None:
                raise ValueError("Unsupported formula operation")
            try:
                return self._bounded(operation(left, right))
            except OverflowError as exc:
                raise ValueError("Formula result exceeds the supported numeric range") from exc
        raise ValueError("Unsupported formula operation")


def safe_number_expression(expression: str) -> float:
    if not isinstance(expression, str) or len(expression) > MAX_FORMULA_LENGTH:
        raise ValueError("Formula is empty or too long")
    try:
        tree = ast.parse(expression, mode="eval")
    except (SyntaxError, MemoryError, RecursionError) as exc:
        raise ValueError("Invalid formula syntax") from exc
    return SafeMath().evaluate(tree)


@dataclass
class SheetModel:
    name: str = "Sheet1"
    cells: dict[str, str] = field(default_factory=dict)
    column_widths: dict[str, int] = field(default_factory=dict)
    cell_formats: dict[str, dict[str, Any]] = field(default_factory=dict)
    comments: dict[str, str] = field(default_factory=dict)
    validations: dict[str, dict[str, Any]] = field(default_factory=dict)
    freeze_panes: str | None = None
    extra: dict[str, Any] = field(default_factory=dict, repr=False, compare=False)

    def __post_init__(self) -> None:
        normalized: dict[str, str] = {}
        for address, value in dict(self.cells).items():
            normalized[normalize_cell(str(address))] = str(value)
        self.cells = normalized
        self.column_widths = {
            str(key).upper(): max(35, min(500, int(value)))
            for key, value in dict(self.column_widths).items()
            if str(key).isalpha() and column_index(str(key)) < MAX_COLS
        }
        self.cell_formats = {
            normalize_cell(str(address)): dict(settings)
            for address, settings in dict(self.cell_formats).items()
            if isinstance(settings, dict)
        }
        self.comments = {
            normalize_cell(str(address)): str(comment)
            for address, comment in dict(self.comments).items()
            if str(comment)
        }
        self.validations = {
            normalize_cell(str(address)): dict(rule)
            for address, rule in dict(self.validations).items()
            if isinstance(rule, dict)
        }
        if self.freeze_panes is not None:
            self.freeze_panes = normalize_cell(self.freeze_panes)

    @classmethod
    def from_dict(cls, payload: dict[str, Any]) -> "SheetModel":
        if not isinstance(payload, dict):
            raise DataCorruptionError("Invalid worksheet data.")
        known = {"name", "cells", "column_widths", "cell_formats", "comments", "validations", "freeze_panes"}
        name = payload.get("name", "Sheet1")
        cells = payload.get("cells", {})
        widths = payload.get("column_widths", {})
        formats = payload.get("cell_formats", {})
        comments = payload.get("comments", {})
        validations = payload.get("validations", {})
        freeze_panes = payload.get("freeze_panes")
        if not isinstance(name, str) or not name or len(name) > 31:
            raise DataCorruptionError("Invalid worksheet name.")
        if not isinstance(cells, dict) or not isinstance(widths, dict):
            raise DataCorruptionError("Invalid worksheet cells or widths.")
        if not isinstance(formats, dict) or not isinstance(comments, dict) or not isinstance(validations, dict):
            raise DataCorruptionError("Invalid worksheet formatting, comments, or validation data.")
        if freeze_panes is not None and not isinstance(freeze_panes, str):
            raise DataCorruptionError("Invalid worksheet freeze pane.")

        normalized_cells: dict[str, str] = {}
        for address, value in cells.items():
            if not isinstance(address, str) or not isinstance(value, str):
                raise DataCorruptionError("Worksheet cell addresses and values must be text.")
            try:
                normalized_cells[normalize_cell(address)] = value
            except (TypeError, ValueError) as exc:
                raise DataCorruptionError(f"Invalid worksheet cell address: {address!r}.") from exc

        normalized_widths: dict[str, int] = {}
        for column, width in widths.items():
            if not isinstance(column, str) or not column.isalpha():
                raise DataCorruptionError("Worksheet column width keys must be column names.")
            if isinstance(width, bool) or not isinstance(width, int) or not 35 <= width <= 500:
                raise DataCorruptionError("Worksheet column widths must be integers from 35 through 500.")
            upper = column.upper()
            if column_index(upper) >= MAX_COLS:
                raise DataCorruptionError(f"Worksheet column {upper!r} is outside the supported grid.")
            normalized_widths[upper] = width

        return cls(
            name=name,
            cells=normalized_cells,
            column_widths=normalized_widths,
            cell_formats=formats,
            comments=comments,
            validations=validations,
            freeze_panes=freeze_panes,
            extra={k: v for k, v in payload.items() if k not in known},
        )

    def to_dict(self) -> dict[str, Any]:
        return merge_known_and_extra(
            {
                "name": self.name,
                "cells": dict(self.cells),
                "column_widths": dict(self.column_widths),
                "cell_formats": dict(self.cell_formats),
                "comments": dict(self.comments),
                "validations": dict(self.validations),
                "freeze_panes": self.freeze_panes,
            },
            self.extra,
        )

    def raw(self, address: str) -> str:
        return self.cells.get(normalize_cell(address), "")

    def set(self, address: str, value: str) -> None:
        key = normalize_cell(address)
        if value == "":
            self.cells.pop(key, None)
        else:
            self.cells[key] = str(value)

    def set_format(self, address: str, **settings: Any) -> None:
        key = normalize_cell(address)
        allowed = {"bold", "italic", "underline", "fill", "foreground", "align", "number_format"}
        current = dict(self.cell_formats.get(key, {}))
        current.update({name: value for name, value in settings.items() if name in allowed})
        current = {name: value for name, value in current.items() if value not in (None, False, "")}
        if current:
            self.cell_formats[key] = current
        else:
            self.cell_formats.pop(key, None)

    def display_value(self, address: str) -> str:
        value = self.value(address)
        number_format = self.cell_formats.get(normalize_cell(address), {}).get("number_format")
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            if number_format == "percent":
                return f"{value * 100:.2f}%"
            if number_format == "currency":
                return f"${value:,.2f}"
            if number_format == "integer":
                return f"{value:,.0f}"
        return str(value)

    def _shift_cells(self, *, axis: int, index: int, delta: int, delete_count: int = 0) -> None:
        def shifted(source: dict[str, Any]) -> dict[str, Any]:
            result: dict[str, Any] = {}
            for address, value in source.items():
                row, col = split_cell(address)
                coordinate = row if axis == 0 else col
                if delete_count and index <= coordinate < index + delete_count:
                    continue
                if coordinate >= index + delete_count:
                    coordinate += delta
                if axis == 0:
                    row = coordinate
                else:
                    col = coordinate
                if 0 <= row < MAX_ROWS and 0 <= col < MAX_COLS:
                    result[f"{column_name(col)}{row + 1}"] = value
            return result

        self.cells = shifted(self.cells)
        self.cell_formats = shifted(self.cell_formats)
        self.comments = shifted(self.comments)
        self.validations = shifted(self.validations)

    def insert_rows(self, index: int, count: int = 1) -> None:
        if not 0 <= index < MAX_ROWS or count < 1:
            raise ValueError("Invalid row insertion")
        self._shift_cells(axis=0, index=index, delta=count)

    def delete_rows(self, index: int, count: int = 1) -> None:
        if not 0 <= index < MAX_ROWS or count < 1:
            raise ValueError("Invalid row deletion")
        self._shift_cells(axis=0, index=index, delta=-count, delete_count=count)

    def insert_columns(self, index: int, count: int = 1) -> None:
        if not 0 <= index < MAX_COLS or count < 1:
            raise ValueError("Invalid column insertion")
        self._shift_cells(axis=1, index=index, delta=count)

    def delete_columns(self, index: int, count: int = 1) -> None:
        if not 0 <= index < MAX_COLS or count < 1:
            raise ValueError("Invalid column deletion")
        self._shift_cells(axis=1, index=index, delta=-count, delete_count=count)

    def sort_range(self, start: str, end: str, *, key_column: int | None = None, reverse: bool = False) -> None:
        top, left, bottom, right = SelectionRange(start, end).bounds()
        key_column = left if key_column is None else key_column
        if not left <= key_column <= right:
            raise ValueError("Sort key must be inside the selected range")
        rows = [
            [self.raw(f"{column_name(col)}{row + 1}") for col in range(left, right + 1)]
            for row in range(top, bottom + 1)
        ]

        def key(values: list[str]):
            raw = values[key_column - left]
            try:
                return (0, float(raw))
            except ValueError:
                return (1, raw.casefold())

        rows.sort(key=key, reverse=reverse)
        for row_offset, values in enumerate(rows):
            for col_offset, value in enumerate(values):
                self.set(f"{column_name(left + col_offset)}{top + row_offset + 1}", value)

    def value(self, address: str, stack: set[str] | None = None) -> Any:
        address = normalize_cell(address)
        stack = set(stack or ())
        if len(stack) > MAX_FORMULA_DEPTH:
            return "#ERROR!"
        if address in stack:
            return "#CYCLE!"
        stack.add(address)
        raw = self.raw(address)
        if not raw.startswith("="):
            if raw == "":
                return ""
            try:
                return float(raw) if "." in raw else int(raw)
            except ValueError:
                return raw
        try:
            return self.evaluate(raw[1:], stack)
        except ZeroDivisionError:
            return "#DIV/0!"
        except Exception:
            return "#ERROR!"

    def evaluate(self, expression: str, stack: set[str]) -> Any:
        if len(expression) > MAX_FORMULA_LENGTH:
            raise ValueError("Formula is too long")
        working = expression.strip()
        references = 0

        def function_replace(match: re.Match[str]) -> str:
            nonlocal references
            name = match.group(1).upper()
            content = match.group(2).strip()
            values: list[float] = []
            nonempty = 0
            parts = [part.strip() for part in content.split(",") if part.strip()]
            for part in parts:
                range_match = RANGE_RE.fullmatch(part.upper())
                if range_match:
                    addresses = list(iter_range(*range_match.groups()))
                elif CELL_RE.fullmatch(part.upper()):
                    addresses = [part.upper()]
                else:
                    try:
                        values.append(float(part))
                        nonempty += 1
                    except ValueError:
                        if part.strip('"'):
                            nonempty += 1
                    continue
                for address in addresses:
                    references += 1
                    if references > MAX_FORMULA_REFERENCES:
                        raise ValueError("Formula references too many cells")
                    value = self.value(address, stack)
                    if value != "":
                        nonempty += 1
                    if isinstance(value, (int, float)) and not isinstance(value, bool):
                        values.append(float(value))
            if name == "SUM":
                answer = sum(values)
            elif name == "AVERAGE":
                answer = sum(values) / len(values) if values else 0
            elif name == "MIN":
                answer = min(values) if values else 0
            elif name == "MAX":
                answer = max(values) if values else 0
            elif name == "COUNTA":
                answer = nonempty
            elif name == "ABS":
                answer = abs(values[0]) if values else 0
            elif name == "ROUND":
                digits = int(values[1]) if len(values) > 1 else 0
                answer = round(values[0], max(-15, min(15, digits))) if values else 0
            else:
                answer = len(values)
            if not math.isfinite(float(answer)) or abs(float(answer)) > MAX_FORMULA_MAGNITUDE:
                raise ValueError("Formula result exceeds the supported numeric range")
            return str(answer)

        passes = 0
        while FUNCTION_RE.search(working):
            passes += 1
            if passes > MAX_FORMULA_NODES:
                raise ValueError("Formula is too complex")
            working = FUNCTION_RE.sub(function_replace, working)

        def cell_replace(match: re.Match[str]) -> str:
            nonlocal references
            references += 1
            if references > MAX_FORMULA_REFERENCES:
                raise ValueError("Formula references too many cells")
            value = self.value(match.group(1), stack)
            if isinstance(value, str) and value.startswith("#"):
                raise ValueError(value)
            return str(value if isinstance(value, (int, float)) and not isinstance(value, bool) else 0)

        working = CELL_RE.sub(cell_replace, working.upper())
        result = safe_number_expression(working)
        return int(result) if result.is_integer() else round(result, 10)


@dataclass
class WorkbookModel:
    title: str = "Untitled Workbook"
    sheets: list[SheetModel] = field(default_factory=lambda: [SheetModel()])
    format_version: int = WORKBOOK_FORMAT_VERSION
    extra: dict[str, Any] = field(default_factory=dict, repr=False, compare=False)

    @classmethod
    def from_dict(cls, payload: dict[str, Any]) -> "WorkbookModel":
        if not isinstance(payload, dict):
            raise DataCorruptionError("Invalid LeanDesk workbook root.")
        version = payload.get("format_version", WORKBOOK_FORMAT_VERSION)
        if isinstance(version, bool) or not isinstance(version, int) or version < 1:
            raise DataCorruptionError("Invalid workbook format version.")
        if version > WORKBOOK_FORMAT_VERSION:
            raise UnsupportedSchemaVersion(version, WORKBOOK_FORMAT_VERSION)
        title = payload.get("title", "Untitled Workbook")
        rows = payload.get("sheets", [])
        if not isinstance(title, str) or len(title) > 4096:
            raise DataCorruptionError("Invalid workbook title.")
        if not isinstance(rows, list) or len(rows) > 256:
            raise DataCorruptionError("Invalid workbook sheet list.")
        if any(not isinstance(row, dict) for row in rows):
            raise DataCorruptionError("Every workbook sheet must be an object.")
        sheets = [SheetModel.from_dict(row) for row in rows]
        known = {"title", "sheets", "format_version"}
        return cls(
            title=title,
            sheets=sheets or [SheetModel()],
            format_version=version,
            extra={k: v for k, v in payload.items() if k not in known},
        )

    def to_dict(self) -> dict[str, Any]:
        return merge_known_and_extra(
            {
                "format_version": WORKBOOK_FORMAT_VERSION,
                "title": self.title,
                "sheets": [sheet.to_dict() for sheet in self.sheets],
            },
            self.extra,
        )


class SheetGrid(ttk.Frame):
    ROWS = MAX_ROWS
    COLS = MAX_COLS
    ROW_HEIGHT = 29
    HEADER_HEIGHT = 32
    ROW_HEADER_WIDTH = 48

    def __init__(self, master, model: SheetModel, on_change, on_selection, on_reference=None):
        super().__init__(master)
        self.model = model
        self.on_change = on_change
        self.on_selection = on_selection
        self.on_reference = on_reference
        self.column_widths = [model.column_widths.get(column_name(index), 92) for index in range(self.COLS)]
        self.canvas = tk.Canvas(
            self, background=SHEET_CANVAS, highlightthickness=1,
            highlightbackground=SHEET_GRID, takefocus=True,
        )
        # Compatibility alias retained for automation that locates SheetGrid.tree.
        self.tree = self.canvas
        yscroll = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        xscroll = ttk.Scrollbar(self, orient="horizontal", command=self.canvas.xview)
        self.canvas.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        self.canvas.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")
        self.rowconfigure(0, weight=1)
        self.columnconfigure(0, weight=1)
        self.editor: ttk.Entry | None = None
        self.active_address = "A1"
        self.selection = SelectionRange()
        self._drag_selecting = False
        self._undo_stack: list[list[tuple[str, str, str]]] = []
        self._redo_stack: list[list[tuple[str, str, str]]] = []
        self._clipboard_origin: tuple[int, int] | None = None
        self._clipboard_text: str | None = None
        self._resize_column: int | None = None
        self._resize_start_x = 0.0
        self._resize_start_width = 0
        self.canvas.bind("<Button-1>", self._button_press)
        self.canvas.bind("<B1-Motion>", self._drag_motion)
        self.canvas.bind("<ButtonRelease-1>", self._button_release)
        self.canvas.bind("<Double-1>", self.begin_edit)
        self.canvas.bind("<Return>", self.begin_edit)
        self.canvas.bind("<Up>", lambda event: self._move_selection(-1, 0))
        self.canvas.bind("<Down>", lambda event: self._move_selection(1, 0))
        self.canvas.bind("<Left>", lambda event: self._move_selection(0, -1))
        self.canvas.bind("<Right>", lambda event: self._move_selection(0, 1))
        self.canvas.bind("<Shift-Up>", lambda event: self._move_selection(-1, 0, extend=True))
        self.canvas.bind("<Shift-Down>", lambda event: self._move_selection(1, 0, extend=True))
        self.canvas.bind("<Shift-Left>", lambda event: self._move_selection(0, -1, extend=True))
        self.canvas.bind("<Shift-Right>", lambda event: self._move_selection(0, 1, extend=True))
        self.canvas.bind("<Control-a>", self.select_all)
        self.canvas.bind("<Control-c>", self.copy_selection)
        self.canvas.bind("<Control-x>", self.cut_selection)
        self.canvas.bind("<Control-v>", self.paste_selection)
        self.canvas.bind("<Control-z>", self.undo)
        self.canvas.bind("<Control-y>", self.redo)
        self.canvas.bind("<Delete>", self.clear_selection)
        self.canvas.bind("<MouseWheel>", self._mousewheel)
        self.canvas.bind("<Shift-MouseWheel>", self._horizontal_mousewheel)
        self.bind("<<LeanDeskThemeChanged>>", lambda _event: self.refresh())
        self.refresh()

    def refresh(self) -> None:
        self.canvas.delete("grid")
        self._x_positions = [self.ROW_HEADER_WIDTH]
        for width in self.column_widths:
            self._x_positions.append(self._x_positions[-1] + width)
        total_width = self._x_positions[-1]
        total_height = self.HEADER_HEIGHT + self.ROWS * self.ROW_HEIGHT
        self.canvas.configure(scrollregion=(0, 0, total_width, total_height), background=SHEET_CANVAS, highlightbackground=SHEET_GRID)
        self.canvas.create_rectangle(0, 0, self.ROW_HEADER_WIDTH, self.HEADER_HEIGHT, fill=COLORS["panel2"], outline=SHEET_GRID, tags="grid")
        self.canvas.create_text(self.ROW_HEADER_WIDTH / 2, self.HEADER_HEIGHT / 2, text="#", fill=COLORS["text"], font=("Segoe UI Semibold", 9), tags="grid")
        for col in range(self.COLS):
            left, right = self._x_positions[col], self._x_positions[col + 1]
            self.canvas.create_rectangle(left, 0, right, self.HEADER_HEIGHT, fill=COLORS["panel2"], outline=SHEET_GRID, tags="grid")
            self.canvas.create_text((left + right) / 2, self.HEADER_HEIGHT / 2, text=column_name(col), fill=COLORS["text"], font=("Segoe UI Semibold", 9), tags="grid")
        active_row, active_col = split_cell(self.active_address)
        top_row, left_col, bottom_row, right_col = self.selection.bounds()
        for row in range(self.ROWS):
            top = self.HEADER_HEIGHT + row * self.ROW_HEIGHT
            bottom = top + self.ROW_HEIGHT
            row_selected = top_row <= row <= bottom_row and left_col == 0 and right_col == self.COLS - 1
            self.canvas.create_rectangle(0, top, self.ROW_HEADER_WIDTH, bottom, fill=COLORS["selection"] if row_selected or row == active_row else COLORS["panel2"], outline=SHEET_GRID, tags="grid")
            self.canvas.create_text(self.ROW_HEADER_WIDTH / 2, (top + bottom) / 2, text=str(row + 1), fill=COLORS["button_active_text"] if row == active_row else COLORS["text"], font=("Segoe UI", 9), tags="grid")
            for col in range(self.COLS):
                left, right = self._x_positions[col], self._x_positions[col + 1]
                in_range = top_row <= row <= bottom_row and left_col <= col <= right_col
                selected = row == active_row and col == active_col
                fill = COLORS["selection"] if selected else SHEET_RANGE_FILL if in_range else SHEET_CANVAS
                self.canvas.create_rectangle(left, top, right, bottom, fill=fill, outline=COLORS["focus"] if selected else SHEET_GRID, width=2 if selected else 1, tags="grid")
                address = f"{column_name(col)}{row + 1}"
                value = self.model.value(address)
                if value != "":
                    formatting = self.model.cell_formats.get(address, {})
                    if not in_range and formatting.get("fill"):
                        self.canvas.itemconfigure(self.canvas.find_withtag("grid")[-1], fill=formatting["fill"])
                    anchor = {"center": "center", "right": "e"}.get(formatting.get("align"), "w")
                    text_x = (left + right) / 2 if anchor == "center" else right - 6 if anchor == "e" else left + 6
                    weight = "Segoe UI Semibold" if formatting.get("bold") else "Segoe UI"
                    foreground = formatting.get("foreground", SHEET_TEXT)
                    self.canvas.create_text(text_x, (top + bottom) / 2, text=self.model.display_value(address), anchor=anchor, width=max(1, right - left - 12), fill=COLORS["button_active_text"] if selected else foreground, font=(weight, 9, "italic" if formatting.get("italic") else "normal"), tags="grid")

    def _column_at(self, x: float) -> int | None:
        for col in range(self.COLS):
            if self._x_positions[col] <= x < self._x_positions[col + 1]:
                return col
        return None

    def _address_from_event(self, event=None) -> str:
        if event is None:
            return self.active_address
        x = self.canvas.canvasx(event.x)
        y = self.canvas.canvasy(event.y)
        col = self._column_at(x)
        row = int((y - self.HEADER_HEIGHT) // self.ROW_HEIGHT)
        if col is not None and 0 <= row < self.ROWS:
            return f"{column_name(col)}{row + 1}"
        return self.active_address

    def _selection_changed(self, event=None) -> None:
        address = self._address_from_event(event)
        self.select_address(address, extend=bool(getattr(event, "state", 0) & 0x0001))

    def select_address(self, address: str, *, extend: bool = False) -> None:
        row, col = split_cell(address)
        self.active_address = address.upper()
        anchor = self.selection.anchor if extend else self.active_address
        self.selection = SelectionRange(anchor, self.active_address)
        self.on_selection(self.active_address, self.model.raw(self.active_address))
        self.refresh()
        self.update_idletasks()
        left, top, right, bottom = self._cell_bounds(row, col)
        visible_left, visible_top = self.canvas.canvasx(0), self.canvas.canvasy(0)
        visible_right = visible_left + self.canvas.winfo_width()
        visible_bottom = visible_top + self.canvas.winfo_height()
        total_width = max(1, self._x_positions[-1])
        total_height = max(1, self.HEADER_HEIGHT + self.ROWS * self.ROW_HEIGHT)
        if left < visible_left or right > visible_right:
            self.canvas.xview_moveto(max(0.0, min(1.0, (left - self.ROW_HEADER_WIDTH) / total_width)))
        if top < visible_top or bottom > visible_bottom:
            self.canvas.yview_moveto(max(0.0, min(1.0, (top - self.HEADER_HEIGHT) / total_height)))

    def select_range(self, anchor: str, active: str) -> None:
        normalize_cell(anchor)
        normalize_cell(active)
        self.selection = SelectionRange(anchor.upper(), active.upper())
        self.active_address = active.upper()
        self.on_selection(self.active_address, self.model.raw(self.active_address))
        self.refresh()

    def selected_addresses(self) -> list[str]:
        return self.selection.addresses()

    def show_formula_reference(self, anchor: str, active: str) -> None:
        """Outline a reference without moving the formula destination selection."""
        row1, col1, row2, col2 = SelectionRange(anchor, active).bounds()
        x1, y1, _, _ = self._cell_bounds(row1, col1)
        _, _, x2, y2 = self._cell_bounds(row2, col2)
        self.canvas.delete("formula-reference")
        self.canvas.create_rectangle(
            x1 + 1, y1 + 1, x2 - 1, y2 - 1,
            outline="#0078d4", width=3, dash=(4, 2), tags="formula-reference",
        )

    def _cell_bounds(self, row: int, col: int) -> tuple[float, float, float, float]:
        return (self._x_positions[col], self.HEADER_HEIGHT + row * self.ROW_HEIGHT, self._x_positions[col + 1], self.HEADER_HEIGHT + (row + 1) * self.ROW_HEIGHT)

    def begin_edit(self, event=None):
        if event is not None:
            address = self._address_from_event(event)
        else:
            address = self.active_address
        row, col = split_cell(address)
        left, top, right, bottom = self._cell_bounds(row, col)
        x = left - self.canvas.canvasx(0)
        y = top - self.canvas.canvasy(0)
        if x >= self.canvas.winfo_width() or y >= self.canvas.winfo_height() or right <= self.canvas.canvasx(0) or bottom <= self.canvas.canvasy(0):
            return "break"
        if self.editor:
            self.editor.destroy()
        self.active_address = address
        self.on_selection(address, self.model.raw(address))
        self.editor = ttk.Entry(self.canvas)
        self.editor.insert(0, self.model.raw(address))
        self.editor.place(x=x + 1, y=y + 1, width=max(20, right - left - 2), height=max(18, bottom - top - 2))
        self.editor.focus_set()
        self.editor.select_range(0, "end")
        self.editor.bind("<Return>", lambda _e: self.commit_edit(address))
        self.editor.bind("<Tab>", lambda _e: self.commit_edit(address))
        self.editor.bind("<Escape>", lambda _e: self.cancel_edit())
        self.editor.bind("<FocusOut>", lambda _e: self.commit_edit(address))
        return "break"

    def commit_edit(self, address: str):
        if not self.editor:
            return "break"
        value = self.editor.get()
        self.editor.destroy()
        self.editor = None
        self.set_cells([(address, value)])
        self.select_address(address)
        return "break"

    def cancel_edit(self):
        if self.editor:
            self.editor.destroy()
            self.editor = None
        return "break"

    def _button_press(self, event):
        x = self.canvas.canvasx(event.x)
        y = self.canvas.canvasy(event.y)
        if y < self.HEADER_HEIGHT:
            for col in range(self.COLS):
                if abs(x - self._x_positions[col + 1]) <= 5:
                    self._resize_column = col
                    self._resize_start_x = x
                    self._resize_start_width = self.column_widths[col]
                    self.canvas.configure(cursor="sb_h_double_arrow")
                    return "break"
            col = self._column_at(x)
            if col is not None:
                self.select_range(f"{column_name(col)}1", f"{column_name(col)}{self.ROWS}")
                self._drag_selecting = False
                self.canvas.focus_set()
                return "break"
        if x < self.ROW_HEADER_WIDTH and y >= self.HEADER_HEIGHT:
            row = max(0, min(self.ROWS - 1, int((y - self.HEADER_HEIGHT) // self.ROW_HEIGHT)))
            self.select_range(f"A{row + 1}", f"{column_name(self.COLS - 1)}{row + 1}")
            self._drag_selecting = False
            self.canvas.focus_set()
            return "break"
        if x < self.ROW_HEADER_WIDTH and y < self.HEADER_HEIGHT:
            return self.select_all()
        address = self._address_from_event(event)
        if self.on_reference and self.on_reference(address):
            self._reference_anchor = address
            self.show_formula_reference(address, address)
            return "break"
        self.canvas.focus_set()
        extend = bool(getattr(event, "state", 0) & 0x0001)
        self.select_address(address, extend=extend)
        self._drag_selecting = True
        return None

    def _drag_motion(self, event):
        if self._resize_column is None:
            if getattr(self, "_reference_anchor", None) is not None:
                address = self._address_from_event(event)
                reference = SelectionRange(self._reference_anchor, address).label()
                if self.on_reference(reference, extend=True):
                    self.show_formula_reference(self._reference_anchor, address)
                return "break"
            if self._drag_selecting:
                address = self._address_from_event(event)
                if address != self.selection.active:
                    self.select_range(self.selection.anchor, address)
                return "break"
            return None
        x = self.canvas.canvasx(event.x)
        width = max(45, min(500, int(self._resize_start_width + x - self._resize_start_x)))
        self.column_widths[self._resize_column] = width
        self.refresh()
        return "break"

    # Compatibility alias for existing GUI automation.
    def _resize_motion(self, event):
        return self._drag_motion(event)

    def _button_release(self, _event):
        if self._resize_column is not None:
            name = column_name(self._resize_column)
            self.model.column_widths[name] = self.column_widths[self._resize_column]
            self._resize_column = None
            self.canvas.configure(cursor="")
            self.on_change()
            return "break"
        self._drag_selecting = False
        self._reference_anchor = None
        return None

    def _move_selection(self, row_delta: int, col_delta: int, *, extend: bool = False):
        row, col = split_cell(self.active_address)
        row = max(0, min(self.ROWS - 1, row + row_delta))
        col = max(0, min(self.COLS - 1, col + col_delta))
        self.select_address(f"{column_name(col)}{row + 1}", extend=extend)
        return "break"

    def select_all(self, _event=None):
        self.select_range("A1", f"{column_name(self.COLS - 1)}{self.ROWS}")
        return "break"

    def _selection_matrix(self) -> list[list[str]]:
        top, left, bottom, right = self.selection.bounds()
        return [
            [self.model.raw(f"{column_name(col)}{row + 1}") for col in range(left, right + 1)]
            for row in range(top, bottom + 1)
        ]

    def copy_selection(self, _event=None):
        stream = io.StringIO(newline="")
        writer = csv.writer(stream, delimiter="\t", lineterminator="\n")
        writer.writerows(self._selection_matrix())
        text = stream.getvalue().rstrip("\n")
        self.clipboard_clear()
        self.clipboard_append(text)
        top, left, _bottom, _right = self.selection.bounds()
        self._clipboard_origin = (top, left)
        self._clipboard_text = text
        return "break"

    def cut_selection(self, _event=None):
        self.copy_selection()
        return self.clear_selection()

    @staticmethod
    def _translate_formula(value: str, row_delta: int, col_delta: int) -> str:
        if not value.startswith("=") or (row_delta == 0 and col_delta == 0):
            return value

        def replace(match: re.Match[str]) -> str:
            row, col = split_cell(match.group(1))
            row = max(0, min(MAX_ROWS - 1, row + row_delta))
            col = max(0, min(MAX_COLS - 1, col + col_delta))
            return f"{column_name(col)}{row + 1}"

        return CELL_RE.sub(replace, value.upper())

    def paste_selection(self, _event=None):
        try:
            text = self.clipboard_get()
        except tk.TclError:
            return "break"
        rows = list(csv.reader(io.StringIO(text), delimiter="\t"))
        start_row, start_col = split_cell(self.active_address)
        if text == self._clipboard_text and self._clipboard_origin is not None:
            formula_row_delta = start_row - self._clipboard_origin[0]
            formula_col_delta = start_col - self._clipboard_origin[1]
        else:
            formula_row_delta = formula_col_delta = 0
        changes: list[tuple[str, str]] = []
        for row_offset, values in enumerate(rows):
            for col_offset, value in enumerate(values):
                row, col = start_row + row_offset, start_col + col_offset
                if row >= self.ROWS or col >= self.COLS:
                    continue
                translated = self._translate_formula(value, formula_row_delta, formula_col_delta)
                changes.append((f"{column_name(col)}{row + 1}", translated))
        self.set_cells(changes)
        if changes:
            self.select_range(self.active_address, changes[-1][0])
        return "break"

    def clear_selection(self, _event=None):
        self.set_cells([(address, "") for address in self.selected_addresses()])
        return "break"

    def set_cells(self, changes: list[tuple[str, str]], *, record: bool = True) -> None:
        transaction: list[tuple[str, str, str]] = []
        for address, new_value in changes:
            old_value = self.model.raw(address)
            if old_value != new_value:
                transaction.append((address, old_value, new_value))
        if not transaction:
            return
        for address, _old_value, new_value in transaction:
            self.model.set(address, new_value)
        if record:
            self._undo_stack.append(transaction)
            self._redo_stack.clear()
        self.on_change()
        self.refresh()

    def undo(self, _event=None):
        if not self._undo_stack:
            return "break"
        transaction = self._undo_stack.pop()
        for address, old_value, _new_value in transaction:
            self.model.set(address, old_value)
        self._redo_stack.append(transaction)
        self.on_change()
        self.refresh()
        return "break"

    def redo(self, _event=None):
        if not self._redo_stack:
            return "break"
        transaction = self._redo_stack.pop()
        for address, _old_value, new_value in transaction:
            self.model.set(address, new_value)
        self._undo_stack.append(transaction)
        self.on_change()
        self.refresh()
        return "break"

    def _mousewheel(self, event):
        self.canvas.yview_scroll(-1 if event.delta > 0 else 1, "units")
        return "break"

    def _horizontal_mousewheel(self, event):
        self.canvas.xview_scroll(-1 if event.delta > 0 else 1, "units")
        return "break"


class SheetsFrame(ttk.Frame):
    def __init__(self, master, *, recent: RecentFiles, on_recent_changed=None, on_title_changed=None):
        super().__init__(master)
        self.recent = recent
        self.on_recent_changed = on_recent_changed
        self.on_title_changed = on_title_changed
        self.current_path: Path | None = None
        self.imported_source_path: Path | None = None
        self.recovery = RecoveryStore()
        self.recovery_id = str(uuid.uuid4())
        self.workbook = WorkbookModel()
        self.dirty = False
        self.active_address = tk.StringVar(value="A1")
        self.formula_var = tk.StringVar()
        self.status_var = tk.StringVar(value="Ready")
        self.summary_var = tk.StringVar(value="1 sheet")
        self._build_ui()
        self.rebuild_tabs()

    def _build_ui(self) -> None:
        ribbon = ResponsiveToolbar(self, bg=COLORS["panel"], height=96, highlightbackground=COLORS["line"], highlightthickness=1)
        ribbon.pack(fill="x")
        ribbon.pack_propagate(False)
        for label, command in (
            ("New", self.new_workbook), ("Open", self.open_workbook), ("Save", self.save),
            ("Save As", self.save_as), ("Add Sheet", self.add_sheet), ("Rename", self.rename_sheet),
            ("Delete Sheet", self.delete_sheet), ("Undo", self.undo), ("Redo", self.redo),
            ("Bold", self.toggle_bold), ("Insert Row", self.insert_row), ("Delete Row", self.delete_row),
            ("Sort A-Z", self.sort_ascending), ("Recalculate", self.recalculate), ("Export CSV", self.export_csv),
        ):
            ttk.Button(ribbon, text=label, command=command).pack(side="left", padx=4, pady=18)
        tk.Label(ribbon, text="SHEETS", bg=COLORS["panel"], fg=COLORS["jade"], font=("Segoe UI Bold", 14)).pack(side="right", padx=18)

        formula = tk.Frame(self, bg=COLORS["panel2"], height=38)
        formula.pack(fill="x")
        formula.pack_propagate(False)
        ttk.Entry(formula, textvariable=self.active_address, width=9).pack(side="left", padx=(8, 5), pady=6)
        tk.Label(formula, text="fx", bg=COLORS["panel2"], fg=COLORS["jade"], font=("Segoe UI Bold", 10)).pack(side="left", padx=5)
        self.formula_entry = ttk.Entry(formula, textvariable=self.formula_var)
        self.formula_entry.pack(side="left", fill="x", expand=True, padx=(0, 8), pady=6)
        self.formula_entry.bind("<Return>", self.commit_formula_bar)
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True)
        self.notebook.bind("<<NotebookTabChanged>>", self._tab_changed)
        status = StatusBar(self)
        status.pack(fill="x")
        status.add_left(self.status_var)
        status.add_right(self.summary_var, muted=True)

    def rebuild_tabs(self) -> None:
        for tab in self.notebook.tabs():
            self.notebook.forget(tab)
        self.grids: list[SheetGrid] = []
        for sheet in self.workbook.sheets:
            grid = SheetGrid(self.notebook, sheet, self.mark_dirty, self.cell_selected, self.insert_formula_reference)
            self.grids.append(grid)
            self.notebook.add(grid, text=sheet.name)
        self.summary_var.set(f"{len(self.workbook.sheets)} sheet{'s' if len(self.workbook.sheets) != 1 else ''}")
        self._update_title()

    def active_index(self) -> int:
        try:
            return self.notebook.index(self.notebook.select())
        except tk.TclError:
            return 0

    def active_grid(self) -> SheetGrid:
        return self.grids[self.active_index()]

    def active_sheet(self) -> SheetModel:
        return self.workbook.sheets[self.active_index()]

    def _tab_changed(self, _event=None) -> None:
        if self.grids:
            grid = self.active_grid()
            self.cell_selected(grid.active_address, grid.model.raw(grid.active_address))

    def cell_selected(self, address: str, raw: str) -> None:
        self.active_address.set(address)
        self.formula_var.set(raw)
        value = self.active_sheet().value(address)
        self.status_var.set(f"{address}: {value}")

    def commit_formula_bar(self, _event=None):
        address = self.active_address.get().strip().upper()
        try:
            split_cell(address)
        except ValueError:
            messagebox.showerror("LeanDesk Sheets", "Cell address must look like A1.", parent=self)
            return "break"
        self.active_grid().set_cells([(address, self.formula_var.get())])
        self.active_grid().select_address(address)
        return "break"

    def cancel_formula_reference(self, event=None) -> str:
        grid = self.active_grid()
        self.formula_var.set(grid.model.raw(grid.active_address))
        self._reference_draft = None
        grid._reference_anchor = None
        grid.canvas.delete("formula-reference")
        grid.canvas.focus_set()
        return "break"

    def insert_formula_reference(self, address: str, *, extend: bool = False) -> bool:
        if self.focus_get() != self.formula_entry or not self.formula_var.get().startswith("="):
            return False
        if extend:
            reference_draft = getattr(self, "_reference_draft", None)
            if reference_draft is None:
                return False
            prefix, suffix, previous = reference_draft
            if self.formula_var.get() != previous:
                return False
            updated = prefix + address + suffix
            self.formula_var.set(updated)
            self.formula_entry.icursor(len(prefix) + len(address))
            self._reference_draft = (prefix, suffix, updated)
            return True
        cursor = self.formula_entry.index(tk.INSERT)
        current = self.formula_var.get()
        prefix = current[:cursor]
        suffix = current[cursor:]
        separator = "" if not prefix or prefix[-1] in "=+-*/^(,:" else "+"
        insertion = separator + address
        self.formula_var.set(prefix + insertion + suffix)
        self.formula_entry.icursor(cursor + len(insertion))
        self._reference_draft = (prefix + separator, suffix, self.formula_var.get())
        if not getattr(self, "_reference_escape_bound", False):
            self.formula_entry.bind("<Escape>", self.cancel_formula_reference, add="+")
            self._reference_escape_bound = True
        return True

    def undo(self) -> None:
        self.active_grid().undo()

    def redo(self) -> None:
        self.active_grid().redo()

    def toggle_bold(self) -> None:
        grid = self.active_grid()
        sheet = self.active_sheet()
        make_bold = not all(sheet.cell_formats.get(address, {}).get("bold", False) for address in grid.selected_addresses())
        for address in grid.selected_addresses():
            sheet.set_format(address, bold=make_bold)
        self.mark_dirty()
        grid.refresh()

    def insert_row(self) -> None:
        row, _col = split_cell(self.active_grid().active_address)
        self.active_sheet().insert_rows(row)
        self.mark_dirty()
        self.active_grid().refresh()

    def delete_row(self) -> None:
        row, _col = split_cell(self.active_grid().active_address)
        self.active_sheet().delete_rows(row)
        self.mark_dirty()
        self.active_grid().refresh()

    def sort_ascending(self) -> None:
        grid = self.active_grid()
        self.active_sheet().sort_range(grid.selection.anchor, grid.selection.active)
        self.mark_dirty()
        grid.refresh()

    def mark_dirty(self) -> None:
        self.dirty = True
        self._update_title()
        try:
            self.recovery.save(
                RecoveryRecord(
                    self.recovery_id,
                    "Sheets",
                    self.workbook.title,
                    str(self.current_path or ""),
                    datetime.now().isoformat(timespec="seconds"),
                    self.workbook.to_dict(),
                )
            )
        except Exception:
            self.status_var.set("Recovery copy could not be updated")

    def _update_title(self) -> None:
        title = self.current_path.name if self.current_path else self.workbook.title
        if self.on_title_changed:
            self.on_title_changed(f"Sheets — {title}", self.dirty)

    def confirm_discard(self) -> bool:
        if not self.dirty:
            return True
        answer = messagebox.askyesnocancel("LeanDesk Sheets", "Save workbook changes?", parent=self)
        if answer is None:
            return False
        return self.save() if answer else True

    def new_workbook(self) -> bool:
        if not self.confirm_discard():
            return False
        self.workbook = WorkbookModel()
        self.current_path = None
        self.imported_source_path = None
        self.recovery.delete(self.recovery_id)
        self.recovery_id = str(uuid.uuid4())
        self.dirty = False
        self.rebuild_tabs()
        return True

    def open_workbook(self, path: Path | None = None) -> bool:
        if not self.confirm_discard():
            return False
        if path is None:
            value = filedialog.askopenfilename(parent=self, filetypes=(("LeanDesk workbook", "*.lsheet"), ("Excel/OpenDocument", "*.xlsx *.xls *.xlsm *.xlsb *.ods *.ots"), ("Data files", "*.csv *.tsv *.dif *.dbf"), ("Legacy/other spreadsheets", "*.numbers *.123 *.wk1 *.wk3 *.wk4"), ("All files", "*.*")))
            if not value:
                return False
            path = Path(value)
        try:
            suffix = path.suffix.lower()
            if suffix == ".xlsx":
                self.workbook = self._load_xlsx(path)
            elif suffix == ".csv":
                self.workbook = self._load_csv(path)
            elif suffix == ".tsv":
                self.workbook = self._load_delimited(path, "\t")
            elif suffix in SHEETS_COMPAT:
                converted = convert_with_libreoffice(path, "Sheets")
                self.workbook = self._load_xlsx(converted.as_file(), title=path.stem)
                self.status_var.set(converted.note)
            else:
                payload = strict_json_load_bytes(read_bounded(path, limit=64 * 1024 * 1024))
                self.workbook = WorkbookModel.from_dict(payload)
        except Exception as exc:
            messagebox.showerror("LeanDesk Sheets", f"Could not open workbook.\n\n{exc}", parent=self)
            return False
        self.current_path = path
        self.imported_source_path = imported_source_for("Sheets", path)
        self.recovery.delete(self.recovery_id)
        self.recovery_id = str(uuid.uuid4())
        self.dirty = False
        self.rebuild_tabs()
        self.recent.add(path, "Sheets")
        if self.on_recent_changed:
            self.on_recent_changed()
        return True

    def _protected_import_source(self) -> Path | None:
        explicit = getattr(self, "imported_source_path", None)
        if explicit is not None:
            return Path(explicit)
        return imported_source_for("Sheets", getattr(self, "current_path", None))

    @mark_save_boundary
    def save(self) -> bool:
        if self._protected_import_source() is not None:
            try:
                proceed = messagebox.askyesno(
                    "Original File Protected",
                    "This workbook was imported from another format. LeanDesk will not overwrite the original because unsupported content could be lost.\n\nSave a new copy instead?",
                    parent=self,
                )
            except Exception:
                return False
            return self.save_as() if proceed else False
        return self.save_as() if self.current_path is None else self._write(self.current_path)

    @mark_save_boundary
    def save_as(self) -> bool:
        value = filedialog.asksaveasfilename(parent=self, defaultextension=".lsheet", filetypes=(("LeanDesk workbook", "*.lsheet"), ("Excel workbook", "*.xlsx"), ("CSV active sheet", "*.csv")))
        if not value:
            return False
        return self._write(Path(value))

    @mark_save_boundary
    def _write(self, path: Path) -> bool:
        try:
            destination = validate_destination(
                "Sheets", path, imported_source=self._protected_import_source(), allow_export_only=True
            )
            suffix = destination.suffix.lower()

            def produce(temporary: Path) -> None:
                if suffix == ".xlsx":
                    self._save_xlsx(temporary)
                elif suffix == ".csv":
                    self._save_csv(temporary)
                else:
                    atomic_write_json(temporary, self.workbook.to_dict())

            write_atomically(destination, produce)
        except (ImportedSourceProtectionError, UnsupportedSaveFormatError, SavePolicyError) as exc:
            messagebox.showwarning("LeanDesk Sheets", str(exc), parent=self)
            return False
        except Exception as exc:
            messagebox.showerror("LeanDesk Sheets", f"Could not save workbook.\n\n{exc}", parent=self)
            return False
        if suffix != ".csv":
            self.current_path = destination
            self.imported_source_path = None
            self.dirty = False
            self.recovery.delete(self.recovery_id)
            self.recent.add(destination, "Sheets")
            if self.on_recent_changed:
                self.on_recent_changed()
            self._update_title()
            self.status_var.set(f"Saved {destination.name}")
        else:
            self.status_var.set(f"Exported {destination.name}; workbook source unchanged")
        return True

    def add_sheet(self) -> None:
        existing = {sheet.name for sheet in self.workbook.sheets}
        index = 1
        while f"Sheet{index}" in existing:
            index += 1
        self.workbook.sheets.append(SheetModel(f"Sheet{index}"))
        self.mark_dirty()
        self.rebuild_tabs()
        self.notebook.select(len(self.workbook.sheets) - 1)

    def rename_sheet(self) -> None:
        sheet = self.active_sheet()
        selected_tab = self.notebook.select()
        value = simpledialog.askstring("Rename sheet", "Sheet name:", initialvalue=sheet.name, parent=self)
        if value and value.strip():
            sheet.name = value.strip()[:31]
            self.mark_dirty()
            # Renaming does not change the grid or the active sheet. Rebuilding
            # every tab discarded selection and silently activated Sheet1.
            self.notebook.tab(selected_tab, text=sheet.name)

    def delete_sheet(self) -> None:
        if len(self.workbook.sheets) == 1:
            messagebox.showinfo("LeanDesk Sheets", "A workbook must keep at least one sheet.", parent=self)
            return
        index = self.active_index()
        if messagebox.askyesno("LeanDesk Sheets", f'Delete "{self.workbook.sheets[index].name}"?', parent=self):
            self.workbook.sheets.pop(index)
            self.mark_dirty()
            self.rebuild_tabs()

    def recalculate(self) -> None:
        for grid in self.grids:
            grid.refresh()
        self.status_var.set("Workbook recalculated")

    def export_csv(self) -> bool:
        value = filedialog.asksaveasfilename(parent=self, defaultextension=".csv", filetypes=(("CSV", "*.csv"),))
        return bool(value) and self._write(Path(value))

    def _save_csv(self, path: Path) -> None:
        sheet = self.active_sheet()
        max_row = max((split_cell(key)[0] for key in sheet.cells), default=0)
        max_col = max((split_cell(key)[1] for key in sheet.cells), default=0)
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            for row in range(max_row + 1):
                writer.writerow([sheet.raw(f"{column_name(col)}{row + 1}") for col in range(max_col + 1)])

    @staticmethod
    def _load_delimited(path: Path, delimiter: str) -> WorkbookModel:
        sheet = SheetModel(path.stem)
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            for row_index, row in enumerate(csv.reader(handle, delimiter=delimiter), 1):
                for col_index, value in enumerate(row):
                    if value:
                        sheet.set(f"{column_name(col_index)}{row_index}", value)
        return WorkbookModel(path.stem, [sheet])

    @staticmethod
    def _load_csv(path: Path) -> WorkbookModel:
        sheet = SheetModel(path.stem)
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            for row_index, row in enumerate(csv.reader(handle), 1):
                for col_index, value in enumerate(row):
                    if value:
                        sheet.set(f"{column_name(col_index)}{row_index}", value)
        return WorkbookModel(path.stem, [sheet])

    def _save_xlsx(self, path: Path) -> None:
        try:
            from openpyxl import Workbook
            from openpyxl.comments import Comment
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as exc:
            raise RuntimeError("XLSX support requires openpyxl.") from exc
        workbook = Workbook()
        workbook.remove(workbook.active)
        for model in self.workbook.sheets:
            sheet = workbook.create_sheet(model.name[:31])
            sheet.freeze_panes = model.freeze_panes
            for column, width in model.column_widths.items():
                sheet.column_dimensions[column].width = max(1, width / 7)
            for address, value in model.cells.items():
                sheet[address] = value
            for address, settings in model.cell_formats.items():
                cell = sheet[address]
                cell.font = Font(
                    bold=bool(settings.get("bold")), italic=bool(settings.get("italic")),
                    underline="single" if settings.get("underline") else None,
                    color=str(settings.get("foreground", "172033")).lstrip("#"),
                )
                if settings.get("fill"):
                    color = str(settings["fill"]).lstrip("#")
                    cell.fill = PatternFill("solid", fgColor=color)
                cell.alignment = Alignment(horizontal=settings.get("align"))
                number_format = settings.get("number_format")
                if number_format == "percent":
                    cell.number_format = "0.00%"
                elif number_format == "currency":
                    cell.number_format = "$#,##0.00"
                elif number_format == "integer":
                    cell.number_format = "#,##0"
            for address, comment in model.comments.items():
                sheet[address].comment = Comment(comment, "LeanDesk")
        workbook.save(path)

    @staticmethod
    def _load_xlsx(path, *, title: str | None = None) -> WorkbookModel:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("XLSX support requires openpyxl.") from exc
        from .ooxml_preflight import prepare_ooxml
        prepared = prepare_ooxml(path, "xlsx")
        workbook = load_workbook(prepared.open(), data_only=False, read_only=False)
        sheets: list[SheetModel] = []
        for source in workbook.worksheets:
            model = SheetModel(source.title)
            model.freeze_panes = str(source.freeze_panes) if source.freeze_panes else None
            for column, dimension in source.column_dimensions.items():
                if dimension.width:
                    model.column_widths[column] = max(35, min(500, int(dimension.width * 7)))
            for row in source.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        model.set(cell.coordinate, str(cell.value))
                    settings: dict[str, Any] = {}
                    if cell.font.bold:
                        settings["bold"] = True
                    if cell.font.italic:
                        settings["italic"] = True
                    if cell.font.underline:
                        settings["underline"] = True
                    if cell.alignment.horizontal in {"left", "center", "right"}:
                        settings["align"] = cell.alignment.horizontal
                    if cell.fill.fill_type == "solid" and cell.fill.fgColor.rgb:
                        settings["fill"] = f"#{cell.fill.fgColor.rgb[-6:]}"
                    if cell.number_format == "0.00%":
                        settings["number_format"] = "percent"
                    elif "$" in cell.number_format:
                        settings["number_format"] = "currency"
                    if settings:
                        model.cell_formats[cell.coordinate] = settings
                    if cell.comment:
                        model.comments[cell.coordinate] = cell.comment.text
            sheets.append(model)
        source_title = title or (Path(path).stem if isinstance(path, (str, Path)) else "Imported Workbook")
        return WorkbookModel(source_title, sheets or [SheetModel()])

    def recover_record(self, record: RecoveryRecord) -> None:
        if record.module != "Sheets" or not self.confirm_discard():
            return
        self.workbook = WorkbookModel.from_dict(record.payload)
        self.current_path = Path(record.original_path) if record.original_path else None
        self.imported_source_path = imported_source_for("Sheets", self.current_path)
        self.recovery_id = record.recovery_id
        self.dirty = True
        self.rebuild_tabs()
        self._update_title()
from .spreadsheet_features import install_sheets_frame_feature_ui, install_workbook_feature_support

install_workbook_feature_support(WorkbookModel)
install_sheets_frame_feature_ui(SheetsFrame)
